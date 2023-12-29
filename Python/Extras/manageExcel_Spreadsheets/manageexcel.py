import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
print(type(wb))

# Get sheetnames
print(wb.sheetnames)

# Get a sheet and its title
sheet = wb['Sheet1']
print(sheet)
print("Sheet Title : " + sheet.title)

# Get sheet max rows and columns
print("Sheet Max rows : ", sheet.max_row)
print("Sheet Max cols : ", sheet.max_column)

# Get activesheet
activesheet = wb.active
print("Active Sheet : " + str(activesheet))

# Get sheet data
print(sheet['A1'].value)
print(sheet['B1'].value)

c = sheet['B1']

# Print row, column and value from a cell with coordinates
print('Row : {}, Column : {}, Value : {}, from a cell with coordinates {}'.format(c.row, c.column, c.value, c.coordinate))

# Get cell value using row and column analogy
print(sheet.cell(row = 1, column = 2).value)

for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)

